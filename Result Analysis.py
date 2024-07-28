import fitz  # PyMuPDF
import re
import os
import pymysql.cursors
import pandas as pd  # Import pandas library for handling dataframes
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
import warnings
from openpyxl.styles import Border, Side
# Connect to MySQL database
connection = pymysql.connect(
    host='localhost',
    user='root',
    password='root',
    database='project',
    cursorclass=pymysql.cursors.DictCursor
)

def create_tables():
    with connection.cursor() as cursor:
        # Create table for extracted data
        sql_create_extracted_data_table = """
        CREATE TABLE IF NOT EXISTS extracted_data (
            id INT AUTO_INCREMENT PRIMARY KEY,
            Usn VARCHAR(255),
            Semester VARCHAR(255),
            StudentName VARCHAR(255),
            SubjectCode VARCHAR(255),
            SubjectName VARCHAR(255),
            InternalMarks INT,
            ExternalMarks INT,
            TotalMarks INT,
            Result VARCHAR(255),
            AnnouncedDate DATE,
            INDEX idx_usn (Usn),
            INDEX idx_semester (Semester),
            INDEX idx_student_name (StudentName),
            INDEX idx_subject_code (SubjectCode),
            INDEX idx_subject_name (SubjectName)
        )
        """
        cursor.execute(sql_create_extracted_data_table)

        # Create table for subject details
        sql_create_subject_details_table = """
        CREATE TABLE IF NOT EXISTS subject_details (
            id INT AUTO_INCREMENT PRIMARY KEY,
            SubjectCode VARCHAR(255),
            SubjectName VARCHAR(255),
            Credits INT
        )
        """
        cursor.execute(sql_create_subject_details_table)
        
        # Create table for SGPA data
        sql_create_sgpa_table = """
        CREATE TABLE IF NOT EXISTS sgpa_data (
            id INT AUTO_INCREMENT PRIMARY KEY,
            Usn VARCHAR(255),
            Semester VARCHAR(255),
            StudentName VARCHAR(255),
            SGPA FLOAT,
            INDEX idx_sgpa (SGPA),  -- Add index for the SGPA column
            FOREIGN KEY (Usn) REFERENCES extracted_data(Usn),
            FOREIGN KEY (Semester) REFERENCES extracted_data(Semester),
            FOREIGN KEY (StudentName) REFERENCES extracted_data(StudentName)
        )

        """
        cursor.execute(sql_create_sgpa_table)
        # Create table for storing top performers
        sql_create_topper_table = """
        CREATE TABLE IF NOT EXISTS topper_data (
            id INT AUTO_INCREMENT PRIMARY KEY,
            Usn VARCHAR(255),
            Semester VARCHAR(255),
            StudentName VARCHAR(255),
            SGPA FLOAT,
            INDEX idx_usn (Usn),
            INDEX idx_semester (Semester),
            INDEX idx_student_name (StudentName),
            FOREIGN KEY (Usn) REFERENCES extracted_data(Usn),
            FOREIGN KEY (Semester) REFERENCES extracted_data(Semester),
            FOREIGN KEY (StudentName) REFERENCES extracted_data(StudentName),
            FOREIGN KEY (SGPA) REFERENCES sgpa_data(SGPA)
        )
        """
        cursor.execute(sql_create_topper_table)

def insert_topper_data(data):
    with connection.cursor() as cursor:
        sql_insert_topper = """
        INSERT INTO topper_data (Usn, Semester, StudentName, SGPA) 
        VALUES (%s, %s, %s, %s)
        """
        for index, row in data.iterrows():
            cursor.execute(sql_insert_topper, (row['University Seat Number'], row['Semester'], row['Student Name'], row['SGPA']))
    connection.commit()

def insert_extracted_data(data):
    with connection.cursor() as cursor:
        sql_insert_data = """
        INSERT INTO extracted_data 
        (Usn, Semester, StudentName, SubjectCode, SubjectName, InternalMarks, ExternalMarks, TotalMarks, Result, AnnouncedDate) 
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """
        for row in data:
            cursor.execute(sql_insert_data, (row['Usn'], row['Semester'], row['StudentName'], row['Subject Code'], row['Subject Name'], row['Internal Marks'], row['External Marks'], row['Total Marks'], row['Result'], row['AnnouncedDate']))
    connection.commit()

def insert_subject_details(subject_details):
    with connection.cursor() as cursor:
        sql_insert_details = """
        INSERT INTO subject_details 
        (SubjectCode, SubjectName, Credits) 
        VALUES (%s, %s, %s)
        """
        for subject_code, details in subject_details.items():
            cursor.execute(sql_insert_details, (subject_code, details['SubjectName'], details['Credits']))
    connection.commit()

def fetch_credits_from_scheme(subject_code):
    try:
        with connection.cursor() as cursor:
            sql_fetch_credits = """
            SELECT credits FROM scheme WHERE code = %s
            """
            cursor.execute(sql_fetch_credits, (subject_code,))
            result = cursor.fetchone()
            if result:
                return result['credits']
            else:
                return 0  # Return 0 if subject code not found
    except Exception as e:
        print(f"Error fetching credits from scheme table: {e}")
        return 0

def fetch_credits(subject_code):
    if subject_code not in subject_details:
        # Fetch credits from the scheme table only if not already fetched
        credits = fetch_credits_from_scheme(subject_code)
        subject_details[subject_code] = credits
        print(f"Credits for subject '{subject_code}': {credits}")
    else:
        credits = subject_details[subject_code]
    return credits
        
def calculate_grade_point(total_marks, credits):
    if 90 <= total_marks <= 100:
        return 10
    elif 80 <= total_marks <= 89:
        return 9
    elif 70 <= total_marks <= 79:
        return 8
    elif 60 <= total_marks <= 69:
        return 7
    elif 55 <= total_marks <= 59:
        return 6
    elif 50 <= total_marks <= 54:
        return 5
    elif 40 <= total_marks <= 49:
        return 4
    else:
        return 0

def extract_text_from_pdf(pdf_file_path):
    pdf_document = fitz.open(pdf_file_path)
    extracted_text = ""
    for page_number in range(len(pdf_document)):
        page = pdf_document.load_page(page_number)
        text = page.get_text()
        extracted_text += text
        c= extracted_text.replace(',', '').replace('&', '')
        d= c.replace('/', '')
    return d

def extract_specific_sections(extracted_text):
    patterns = {
        "University Seat Number": r'University\s+Seat\s+Number\s*:\s*(\S+)',
        "Semester": r'Semester\s*:\s*(\S+)',
        "Student Name": r'Student\s+Name\s*:\s*(.+)',
        "Subject Code": r'Subject\s+Code\s*:\s*(\S+)',
        "Subject Name": r'Subject\s+Name\s*:\s*(.+)',
        "Internal Marks": r'Internal\s+Marks\s*:\s*(\d+)',
        "External Marks": r'External\s+Marks\s*:\s*(\d+)',
        "Total Marks": r'Total\s+Marks\s*:\s*(\d+)',
        "Result": r'Result\s*:\s*(\S+)'
    }

    extracted_sections = {}
    for section, pattern in patterns.items():
        match = re.search(pattern, extracted_text, re.IGNORECASE)
        if match:
            extracted_sections[section] = match.group(1).strip()
        else:
            if section == "Semester":
                extracted_sections[section] = ""
            else:
                extracted_sections[section] = "Unknown"
    return extracted_sections

def extract_table_data(extracted_text, extracted_sections):
    pattern =  r'((?:\d{2}[A-Z0-9]+)|(?:21[A-Z0-9]+))\s+([\w\s#]+)\s+(\d+)\s+(\d+)\s+(\d+)\s+([A-Z])\s+(\d{4}-\d{2}-\d{2})'
    matches = re.findall(pattern, extracted_text)
    extracted_data = []
    for match in matches:
        subject_code, subject_name, internal_marks, external_marks, total_marks, result, announced_date = match
        extracted_data.append({
            "Usn": extracted_sections.get("University Seat Number", ""),
            "Semester": extracted_sections.get("Semester", ""),
            "StudentName": extracted_sections.get("Student Name", ""),
            "Subject Code": subject_code.strip(),
            "Subject Name": subject_name.strip(),
            "Internal Marks": int(internal_marks),
            "External Marks": int(external_marks),
            "Total Marks": int(total_marks),
            "Result": result.strip(),
            "AnnouncedDate": announced_date.strip()
        })

    return extracted_data

def insert_sgpa_data(Usn, Semester, StudentName, SGPA):
    with connection.cursor() as cursor:
        sql_insert_sgpa = """
        INSERT INTO sgpa_data (Usn, Semester, StudentName, SGPA) 
        VALUES (%s, %s, %s, %s)
        """
        cursor.execute(sql_insert_sgpa, (Usn, Semester, StudentName, SGPA))
    connection.commit()

def swap_columns(df, col1, col2):
    for i in range(len(df.columns) - 1):
        if df.columns[i] == col1 and df.columns[i + 1] == col2:
            # Swap values
            df.iloc[:, [i, i + 1]] = df.iloc[:, [i + 1, i]]
            # Swap column names
            df.columns.values[i], df.columns.values[i + 1] = df.columns.values[i + 1], df.columns.values[i]

def extract_top_3_toppers(df_sorted):
    return df_sorted.head(3)[['Semester', 'University Seat Number', 'Student Name','SGPA']]

if __name__ == "__main__":
    folder_path = 'C:\\Users\\Fiza Naaz\\Desktop\\Result\\5semAIML'
    semester_data = defaultdict(list)
    subject_details = {}

    try:
        create_tables()

        num_items_in_folder = len([file_name for file_name in os.listdir(folder_path) if file_name.endswith('.pdf')])
        print(f"Number of items in folder: {num_items_in_folder}")

        for file_name in os.listdir(folder_path):
            if file_name.endswith('.pdf'):
                pdf_file_path = os.path.join(folder_path, file_name)
                extracted_text = extract_text_from_pdf(pdf_file_path)
                extracted_sections = extract_specific_sections(extracted_text)
                table_data = extract_table_data(extracted_text, extracted_sections)
                semester = extracted_sections.get("Semester", "Unknown Semester")
                semester_data[semester].extend(table_data)

        for semester, data in semester_data.items():
            insert_extracted_data(data)

            for row in data:
                subject_code = row['Subject Code']
                subject_name = row['Subject Name']
                total_marks = row['Total Marks']
                credits = fetch_credits(subject_code)
                row['Grade Point'] = calculate_grade_point(total_marks, credits)

        df_list = []
        for semester, data in semester_data.items():
            usn_name_groups = defaultdict(list)
            for row in data:
                usn_name_groups[(row['Usn'], row['StudentName'])].append(row)
            for (usn, name), group in usn_name_groups.items():
                row_dict = {'Semester': semester, 'University Seat Number': usn, 'Student Name': name}
                total_credits = 0
                total_grade_points = 0
                total_marks = 0
                has_fail = any(row.get('Result', '') == 'F' for row in group)  # Check if any row has 'Result' as 'F'
                for row in group:
                    subject_code = row['Subject Code']
                    credits = fetch_credits(subject_code)
                    grade_points = row['Grade Point']
                    # Add credit column next to Total Marks
                    row['Credit'] = credits
                    # Calculate C*GP and set to 0 if the result is 'F'
                    if row.get('Result', '') == 'F':
                        row['C*GP'] = 0
                    else:
                        row['C*GP'] = credits * grade_points
                    total_credits += credits
                    total_grade_points += row['C*GP']  # Use row['C*GP'] to ensure it is 0 for F grades
                    total_marks += row['Total Marks']
                sgpa = total_grade_points / total_credits if total_credits != 0 else 0
                rounded_sgpa = round(sgpa, 2)
                insert_sgpa_data(usn, semester, name, rounded_sgpa)
        
                for i, row in enumerate(group):
                    for key, value in row.items():
                        if key not in ('Semester', 'Usn', 'StudentName'):
                            row_dict[f'{key} {i+1}'] = value
        
                row_dict['Total C*GP'] = total_credits * rounded_sgpa
                row_dict['Total Credits'] = total_credits
                row_dict['SGPA'] = rounded_sgpa
                row_dict['Total Marks'] = total_marks
        
                # Calculate percentage and determine result
                percentage = rounded_sgpa * 10
                if has_fail:  # If any "F" grade in individual "Result" columns, assign "Fail"
                    result = 'Fail'
                elif percentage >= 70:
                    result = 'First Class with Distinction (FCD)'
                elif percentage >= 60:
                    result = 'First Class (FC)'
                elif percentage >= 35:
                    result = 'Second Class (SC)'
                else:
                    result = 'Fail'
                row_dict['%'] = f'{percentage:.2f}%'
                row_dict['Result'] = result
        
                df_list.append(row_dict)
        
        df = pd.DataFrame(df_list)
        
        # Drop all columns named 'AnnouncedDate'
        columns_to_drop = [col for col in df.columns if 'AnnouncedDate' in col]
        df.drop(columns=columns_to_drop, inplace=True)
        
        # Rename columns ending with numbers
        rename_dict = {}
        for column in df.columns:
            # Check if column name ends with a number
            if column[-1].isdigit():
                # Extract the part before the number
                original_column_name = column.rsplit(' ', 1)[0]
                # Add renamed column name to the dictionary
                rename_dict[column] = original_column_name
        
        # Rename columns in the DataFrame
        df.rename(columns=rename_dict, inplace=True)
        
        # Call the function to swap the desired columns
        swap_columns(df, 'Result', 'Grade Point')
        swap_columns(df, 'Result', 'Credit')
        swap_columns(df, 'Result', 'C*GP')
        
        # Write DataFrame to Excel
        excel_file_path = 'C:\\Users\\Fiza Naaz\\Desktop\\Result\\Excel\\5semAIML.xlsx'
        with pd.ExcelWriter(excel_file_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Main', index=False)
        
            # Bold header row and add borders (assuming you have defined the worksheet object somewhere)
            worksheet = writer.sheets['Main']
            header_font = Font(bold=True)
            for cell in worksheet[1]:
                cell.font = header_font
        
            border = Border(left=Side(style='thin'), 
                            right=Side(style='thin'), 
                            top=Side(style='thin'), 
                            bottom=Side(style='thin'))
            for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
                for cell in row:
                    cell.border = border
        
        num_rows_excel_sheet = len(df)
        print(f"Number of rows in Excel sheet: {num_rows_excel_sheet}")
        
        if num_items_in_folder == num_rows_excel_sheet:
            df_sorted = df.sort_values(by='SGPA', ascending=False)
            top_3_toppers_data = extract_top_3_toppers(df_sorted)
            print(top_3_toppers_data)
        
            insert_topper_data(top_3_toppers_data)
        
            with pd.ExcelWriter(excel_file_path, engine='openpyxl', mode='a') as writer:
                if 'Top 3 Toppers' in writer.book.sheetnames:
                    idx = writer.book.sheetnames.index('Top 3 Toppers')
                    writer.book.remove(writer.book.worksheets[idx])
                top_3_toppers_data.to_excel(writer, sheet_name='Top 3 Toppers', index=False)
        
            print("Excel file created successfully.")
            print("Data extraction and storage complete.")
        else:
            print("Number of items in the folder does not match the number of rows in the Excel sheet.")

    except Exception as e:
        print(f"An error occurred: {e}")
